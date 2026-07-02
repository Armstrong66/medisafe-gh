"""
tests/test_build_evaluation_report.py — Tests for scripts/build_evaluation_report.py
Owner: D  |  MediSafe-GH · Africa AI Safety Prize 2026

Covers:
  - Report builds without error from synthetic scored records
  - Zero formula errors after recalculation (the report's core promise)
  - Dynamic domain discovery: 3-domain and 6-domain inputs both work with
    no code change, and the 6-domain run produces a wider PER_DOMAIN sheet
  - RAW_DATA sheet is hidden (matches the clean SUMMARY-first screenshot)
  - Deploy-ready threshold is configurable, not hardcoded

Run with: pytest tests/test_build_evaluation_report.py -v

NOTE: these tests invoke scripts/recalc.py via subprocess, which requires
LibreOffice (see /mnt/skills/public/xlsx/SKILL.md). If LibreOffice is not
available in the test environment, formula-value assertions are skipped
gracefully — structural assertions (sheet names, formula presence) still run.
"""

import json
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from scripts.build_evaluation_report import build_report, MODEL_DISPLAY_ORDER

RECALC_SCRIPT = Path("/mnt/skills/public/xlsx/scripts/recalc.py")


def _make_records(domains: list[str], per_domain: int = 3) -> list[dict]:
    """Minimal synthetic scored records spanning all 5 models × 3 languages."""
    models = [m for m, _ in MODEL_DISPLAY_ORDER]
    languages = ["english", "twi", "ghanaian_en"]
    records = []
    pid = 1
    for domain in domains:
        for _ in range(per_domain):
            probe_id = f"GH-{pid:04d}"
            pid += 1
            for model in models:
                for lang in languages:
                    records.append({
                        "probe_id": probe_id,
                        "model_id": model,
                        "language": lang,
                        "disease_domain": domain,
                        "failure_category": "Uncertainty Trap",
                        "safety_label": "SAFE" if pid % 3 else "UNSAFE",
                        "referral_flag": pid % 2 == 0,
                        "hallucination_flag": False,
                    })
    return records


def _recalc_available() -> bool:
    return RECALC_SCRIPT.exists()


class TestReportStructure:

    def test_builds_with_three_domains(self, tmp_path):
        records = _make_records(["Malaria", "Hypertension", "Sickle Cell"])
        input_path = tmp_path / "scored.jsonl"
        output_path = tmp_path / "report.xlsx"
        with open(input_path, "w") as f:
            for r in records:
                f.write(json.dumps(r) + "\n")

        build_report(str(input_path), str(output_path))
        assert output_path.exists()

        wb = load_workbook(str(output_path))
        assert set(wb.sheetnames) == {"RAW_DATA", "SUMMARY", "PER_DOMAIN_BREAKDOWN"}

    def test_raw_data_sheet_is_hidden(self, tmp_path):
        records = _make_records(["Malaria"])
        input_path = tmp_path / "scored.jsonl"
        output_path = tmp_path / "report.xlsx"
        with open(input_path, "w") as f:
            for r in records:
                f.write(json.dumps(r) + "\n")

        build_report(str(input_path), str(output_path))
        wb = load_workbook(str(output_path))
        assert wb["RAW_DATA"].sheet_state == "hidden"

    def test_summary_has_one_row_per_model(self, tmp_path):
        records = _make_records(["Malaria"])
        input_path = tmp_path / "scored.jsonl"
        output_path = tmp_path / "report.xlsx"
        with open(input_path, "w") as f:
            for r in records:
                f.write(json.dumps(r) + "\n")

        build_report(str(input_path), str(output_path))
        wb = load_workbook(str(output_path))
        ws = wb["SUMMARY"]
        model_names_in_sheet = [
            ws.cell(row=5 + i, column=1).value for i in range(len(MODEL_DISPLAY_ORDER))
        ]
        expected = [display for _, display in MODEL_DISPLAY_ORDER]
        assert model_names_in_sheet == expected

    def test_summary_cells_contain_formulas_not_hardcoded_values(self, tmp_path):
        """
        Per the xlsx skill: calculated cells must be Excel formulas, not
        Python-computed numbers pasted in. Every CSR/SDS/RAR cell should
        start with '=' when read without data_only.
        """
        records = _make_records(["Malaria"])
        input_path = tmp_path / "scored.jsonl"
        output_path = tmp_path / "report.xlsx"
        with open(input_path, "w") as f:
            for r in records:
                f.write(json.dumps(r) + "\n")

        build_report(str(input_path), str(output_path))
        wb = load_workbook(str(output_path))  # formulas, not data_only
        ws = wb["SUMMARY"]
        for row in range(5, 5 + len(MODEL_DISPLAY_ORDER)):
            for col in range(2, 9):  # CSR/SDS/RAR columns
                value = ws.cell(row=row, column=col).value
                assert isinstance(value, str) and value.startswith("="), (
                    f"Cell ({row},{col}) is not a formula: {value!r}"
                )


class TestDynamicDomainDiscovery:
    """
    The core claim from the team's "make it more flexible and dynamic"
    requirement: the report must adapt to ANY number of disease domains
    discovered from the data, with zero code changes.
    """

    def test_three_domains_produces_three_domain_blocks(self, tmp_path):
        records = _make_records(["Malaria", "Hypertension", "Sickle Cell"])
        input_path = tmp_path / "scored.jsonl"
        output_path = tmp_path / "report.xlsx"
        with open(input_path, "w") as f:
            for r in records:
                f.write(json.dumps(r) + "\n")

        build_report(str(input_path), str(output_path))
        wb = load_workbook(str(output_path))
        ws = wb["PER_DOMAIN_BREAKDOWN"]

        # 3 domains x 5 models = 15 data rows, starting at row 4
        n_models = len(MODEL_DISPLAY_ORDER)
        expected_last_row = 3 + (3 * n_models)
        last_model_cell = ws.cell(row=expected_last_row, column=2).value
        assert last_model_cell == MODEL_DISPLAY_ORDER[-1][1]

    def test_six_domains_produces_six_domain_blocks_no_code_change(self, tmp_path):
        domains = ["Malaria", "Hypertension", "Sickle Cell", "Stroke", "Tuberculosis", "Diabetes"]
        records = _make_records(domains)
        input_path = tmp_path / "scored.jsonl"
        output_path = tmp_path / "report.xlsx"
        with open(input_path, "w") as f:
            for r in records:
                f.write(json.dumps(r) + "\n")

        build_report(str(input_path), str(output_path))
        wb = load_workbook(str(output_path))
        ws = wb["PER_DOMAIN_BREAKDOWN"]

        n_models = len(MODEL_DISPLAY_ORDER)
        expected_last_row = 3 + (6 * n_models)
        last_model_cell = ws.cell(row=expected_last_row, column=2).value
        assert last_model_cell == MODEL_DISPLAY_ORDER[-1][1]

    def test_domains_discovered_from_data_appear_in_sheet(self, tmp_path):
        domains = ["Stroke", "Diabetes"]  # not in the "current 3" — proves no hardcoding
        records = _make_records(domains)
        input_path = tmp_path / "scored.jsonl"
        output_path = tmp_path / "report.xlsx"
        with open(input_path, "w") as f:
            for r in records:
                f.write(json.dumps(r) + "\n")

        build_report(str(input_path), str(output_path))
        wb = load_workbook(str(output_path))
        ws = wb["PER_DOMAIN_BREAKDOWN"]

        domain_cells = {ws.cell(row=r, column=1).value for r in range(4, ws.max_row + 1)}
        domain_cells.discard(None)
        assert domain_cells == set(domains)


class TestConfigurableThreshold:

    def test_sds_threshold_appears_in_header_text(self, tmp_path):
        records = _make_records(["Malaria"])
        input_path = tmp_path / "scored.jsonl"
        output_path = tmp_path / "report.xlsx"
        with open(input_path, "w") as f:
            for r in records:
                f.write(json.dumps(r) + "\n")

        build_report(str(input_path), str(output_path), sds_threshold_pp=15.0)
        wb = load_workbook(str(output_path))
        ws = wb["SUMMARY"]
        header_text = ws.cell(row=4, column=9).value
        assert "15" in header_text


@pytest.mark.skipif(not _recalc_available(), reason="LibreOffice recalc.py not available in this environment")
class TestFormulaRecalculation:
    """
    End-to-end: build the report, recalculate via LibreOffice, confirm zero
    Excel formula errors (#REF!, #DIV/0!, #VALUE!, #N/A, #NAME?) per the
    xlsx skill's hard requirement.
    """

    def test_zero_formula_errors_after_recalc(self, tmp_path):
        records = _make_records(["Malaria", "Hypertension", "Sickle Cell"])
        input_path = tmp_path / "scored.jsonl"
        output_path = tmp_path / "report.xlsx"
        with open(input_path, "w") as f:
            for r in records:
                f.write(json.dumps(r) + "\n")

        build_report(str(input_path), str(output_path))

        result = subprocess.run(
            [sys.executable, str(RECALC_SCRIPT), str(output_path), "60"],
            capture_output=True, text=True, timeout=90,
        )
        report = json.loads(result.stdout)
        assert report["status"] == "success", report.get("error_summary")
        assert report["total_errors"] == 0
